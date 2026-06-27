#!/usr/bin/env python3
"""
formulas — Excel formula evaluation, workbook calculation, and export utility.

Commands:
    calc      Calculate a workbook, optionally override inputs, render outputs
    build     Export workbook to portable JSON model
    convert   Convert calculated workbook to CSV/JSON files
    parse     Parse and evaluate a standalone formula string
    info      Print workbook structure (sheets, cells, formulas)
    sheets    List sheet names as a JSON array of strings
    serve     Start Flask API server for a workbook
"""

import argparse
import csv
import json
import os
import sys
from pathlib import Path


def _check_formulas():
    """Check if formulas package is installed, exit with guidance if not."""
    try:
        import formulas  # noqa: F401
    except ImportError:
        print(
            "Error: 'formulas' package not installed.\n"
            "Install with: pip install formulas[excel]==1.3.4\n"
            "For CLI support also: pip install click-log",
            file=sys.stderr,
        )
        sys.exit(1)


def _load_model(fpath, circular=False):
    """Load and finish an ExcelModel from a file path."""
    import formulas

    ext = Path(fpath).suffix.lower()
    if ext == '.json':
        with open(fpath) as f:
            data = json.load(f)
        model = formulas.ExcelModel().from_dict(data, assemble=False)
        model.finish(complete=True, circular=circular)
    else:
        model = formulas.ExcelModel().loads(fpath).finish(circular=circular)
    return model


def _extract_value(value):
    """Extract scalar or list value from a Ranges/Solution entry."""
    import schedula as sh
    from formulas.ranges import Ranges

    if isinstance(value, Ranges):
        data = value.value.tolist()
        if len(data) == 1 and len(data[0]) == 1:
            return data[0][0]
        return data
    if value is sh.EMPTY:
        return None
    return value


# ---------------------------------------------------------------------------
# Commands
# ---------------------------------------------------------------------------

def cmd_calc(args):
    """Calculate a workbook with optional input overrides."""
    _check_formulas()
    import formulas
    import schedula as sh

    model = _load_model(args.file, circular=args.circular)

    # Build input overrides
    inputs = {}
    if args.overwrite:
        for assignment in args.overwrite:
            ref, _, val_str = assignment.partition('=')
            ref = ref.strip()
            val_str = val_str.strip()
            # Parse value
            if len(val_str) >= 2 and val_str[0] == val_str[-1] == '"':
                val = val_str[1:-1]
            elif val_str.upper() in ('TRUE', 'FALSE'):
                val = val_str.upper() == 'TRUE'
            else:
                try:
                    val = float(val_str) if '.' in val_str else int(val_str)
                except ValueError:
                    print(f"Error: Cannot parse value '{val_str}'", file=sys.stderr)
                    sys.exit(1)
            inputs[ref] = val

    # Build output refs
    outputs = list(args.out) if args.out else None

    solution = model.calculate(inputs=inputs or {}, outputs=outputs)

    if args.format == 'json':
        result = {}
        for key, val in sorted(solution.items()):
            if isinstance(key, sh.Token):
                continue
            result[key] = _extract_value(val)
        print(json.dumps(result, indent=2, sort_keys=True))
    elif args.format == 'excel':
        out_dir = args.output or '.'
        model.write(dirpath=out_dir, solution=solution)
        print(f"Wrote calculated workbook to {out_dir}/")
    else:
        # Table format
        for key in sorted(solution.keys()):
            if isinstance(key, sh.Token):
                continue
            val = _extract_value(solution[key])
            short_key = key.split('!')[-1] if '!' in key else key
            print(f"  {short_key:>20s}  =  {val}")


def cmd_build(args):
    """Export workbook to JSON model."""
    _check_formulas()
    model = _load_model(args.file, circular=args.circular)
    exported = model.to_dict()

    # Filter to specific refs if --out given
    if args.out:
        filtered = {k: v for k, v in exported.items() if any(k.endswith(ref) for ref in args.out)}
        exported = filtered

    output = json.dumps(exported, indent=2, sort_keys=True)
    if args.output:
        with open(args.output, 'w') as f:
            f.write(output)
            f.write('\n')
        print(f"Wrote JSON model to {args.output}")
    else:
        print(output)


def cmd_convert(args):
    """Convert a calculated workbook to CSV and/or JSON."""
    _check_formulas()
    from openpyxl import load_workbook

    # First calculate and write
    model = _load_model(args.file, circular=args.circular)
    solution = model.calculate()
    calc_dir = args.calc_dir or '/tmp/formulas_calc_tmp'
    os.makedirs(calc_dir, exist_ok=True)
    model.write(dirpath=calc_dir, solution=solution)

    # Find the output workbook (uppercase name)
    src_name = Path(args.file).stem.upper() + '.XLSX'
    src_path = os.path.join(calc_dir, src_name)
    if not os.path.exists(src_path):
        # List what was actually written
        files = os.listdir(calc_dir)
        xlsx_files = [f for f in files if f.endswith('.xlsx') or f.endswith('.XLSX')]
        if xlsx_files:
            src_path = os.path.join(calc_dir, xlsx_files[0])
        else:
            print(f"Error: No calculated workbook found in {calc_dir}", file=sys.stderr)
            sys.exit(1)

    out_dir = args.output or '.'
    os.makedirs(out_dir, exist_ok=True)
    wb = load_workbook(src_path, data_only=True)

    sheets_data = {}
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        # Filter rows that have at least one non-None value
        rows = [list(r) for r in rows if any(v is not None for v in r)]
        if not rows:
            continue

        # CSV export
        if args.csv:
            csv_path = os.path.join(out_dir, f"{ws.title}.csv")
            with open(csv_path, 'w', newline='') as f:
                writer = csv.writer(f)
                for row in rows:
                    writer.writerow(row)
            print(f"  CSV: {csv_path}")

        # JSON export (first row as headers)
        if args.json:
            headers = [v if v is not None else f'col_{i}' for i, v in enumerate(rows[0])]
            data = []
            for row in rows[1:]:
                record = {headers[j]: row[j] for j in range(len(headers))}
                data.append(record)
            json_path = os.path.join(out_dir, f"{ws.title}.json")
            with open(json_path, 'w') as f:
                json.dump(data, f, indent=2)
            print(f"  JSON: {json_path}")

        sheets_data[ws.title] = data if args.json else rows

    # Combined JSON (all sheets)
    if args.combined_json:
        combined_path = os.path.join(out_dir, 'all_sheets.json')
        with open(combined_path, 'w') as f:
            json.dump(sheets_data, f, indent=2, default=str)
        print(f"  Combined JSON: {combined_path}")

    # Cleanup temp dir
    if args.calc_dir is None and os.path.exists(calc_dir):
        import shutil
        shutil.rmtree(calc_dir, ignore_errors=True)


def cmd_parse(args):
    """Parse and evaluate a standalone formula."""
    _check_formulas()
    import formulas

    formula = args.formula
    if not formula.startswith('='):
        formula = '=' + formula

    results = formulas.Parser().ast(formula)
    func = results[1].compile()

    # Show inputs
    input_names = list(func.inputs)
    print(f"Formula: {formula}")
    print(f"Inputs:  {input_names}")

    # Evaluate with provided values or defaults
    if args.values:
        vals = []
        for v in args.values:
            try:
                vals.append(float(v) if '.' in v else int(v))
            except ValueError:
                vals.append(v)
        result = func(*vals)
    else:
        result = func()

    print(f"Result:  {result}")


def cmd_info(args):
    """Print workbook structure info."""
    _check_formulas()
    from openpyxl import load_workbook

    wb = load_workbook(args.file, data_only=False)
    print(f"File: {args.file}")
    print(f"Sheets: {wb.sheetnames}")

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        formula_cells = []
        value_cells = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    if hasattr(cell, 'data_type') and cell.data_type == 'f':
                        formula_cells.append(f"  {cell.coordinate}: {cell.value}")
                    elif not isinstance(cell.value, str) or len(str(cell.value)) < 50:
                        value_cells.append(f"  {cell.coordinate}: {cell.value}")

        print(f"\n--- {ws_name} ---")
        if formula_cells:
            print(f"  Formulas ({len(formula_cells)}):")
            for fc in formula_cells:
                print(fc)
        if value_cells and not args.formulas_only:
            print(f"  Values ({len(value_cells)}):")
            for vc in value_cells[:20]:
                print(vc)
            if len(value_cells) > 20:
                print(f"  ... and {len(value_cells) - 20} more")

    # Also show formulas model info
    model = _load_model(args.file)
    cells = model.cells
    print(f"\nModel nodes: {len(cells)}")
    formula_nodes = [k for k, c in cells.items() if hasattr(c, 'func')]
    print(f"Formula cells: {len(formula_nodes)}")


def cmd_sheets(args):
    """Return sheet names as a JSON array."""
    _check_formulas()
    from openpyxl import load_workbook

    wb = load_workbook(args.file, data_only=False)
    print(json.dumps(wb.sheetnames))


def cmd_serve(args):
    """Start Flask API server."""
    _check_formulas()
    from formulas.app import create_app

    app = create_app(files=(args.file,), circular=args.circular)
    print(f"Serving formulas API on http://{args.host}:{args.port}")
    app.run(host=args.host, port=args.port, debug=False)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        prog='formulas-1-3-4',
        description='Excel formula evaluation, workbook calculation, and export.',
    )
    subparsers = parser.add_subparsers(dest='command', help='Available commands')

    # calc
    p_calc = subparsers.add_parser('calc', help='Calculate a workbook')
    p_calc.add_argument('file', help='Excel (.xlsx/.ods) or JSON model file')
    p_calc.add_argument('--overwrite', '-O', action='append',
                        help="Override cell: REF=VALUE (repeatable)")
    p_calc.add_argument('--out', action='append',
                        help='Output cell/range to compute (repeatable)')
    p_calc.add_argument('--format', '-f', choices=['json', 'excel', 'table'],
                        default='table', help='Output format (default: table)')
    p_calc.add_argument('--output', '-o', help='Output dir (for excel) or file (for json)')
    p_calc.add_argument('--circular', action='store_true',
                        help='Enable circular reference solving')

    # build
    p_build = subparsers.add_parser('build', help='Export to JSON model')
    p_build.add_argument('file', help='Excel (.xlsx/.ods) file')
    p_build.add_argument('--out', action='append',
                         help='Filter to specific cell refs (repeatable)')
    p_build.add_argument('--output', '-o', help='Output JSON file (default: stdout)')
    p_build.add_argument('--circular', action='store_true',
                         help='Enable circular reference solving')

    # convert
    p_conv = subparsers.add_parser('convert', help='Convert to CSV/JSON')
    p_conv.add_argument('file', help='Excel (.xlsx/.ods) file')
    p_conv.add_argument('--csv', action='store_true', help='Export as CSV per sheet')
    p_conv.add_argument('--json', action='store_true', help='Export as JSON per sheet')
    p_conv.add_argument('--combined-json', action='store_true',
                        help='Export all sheets to one JSON file')
    p_conv.add_argument('--output', '-o', default='.', help='Output directory')
    p_conv.add_argument('--calc-dir', help='Temp dir for calculated workbook')
    p_conv.add_argument('--circular', action='store_true',
                        help='Enable circular reference solving')

    # parse
    p_parse = subparsers.add_parser('parse', help='Parse standalone formula')
    p_parse.add_argument('formula', help='Formula string (with or without leading =)')
    p_parse.add_argument('--values', '-v', nargs='*',
                         help='Input values for cell references')

    # info
    p_info = subparsers.add_parser('info', help='Show workbook structure')
    p_info.add_argument('file', help='Excel (.xlsx/.ods) file')
    p_info.add_argument('--formulas-only', action='store_true',
                        help='Only show formula cells')

    # sheets
    p_sheets = subparsers.add_parser('sheets', help='List sheet names as JSON array')
    p_sheets.add_argument('file', help='Excel (.xlsx/.ods) file')

    # serve
    p_serve = subparsers.add_parser('serve', help='Start Flask API server')
    p_serve.add_argument('file', help='Excel (.xlsx/.ods) file')
    p_serve.add_argument('--host', default='127.0.0.1', help='Host (default: 127.0.0.1)')
    p_serve.add_argument('--port', type=int, default=5000, help='Port (default: 5000)')
    p_serve.add_argument('--circular', action='store_true',
                         help='Enable circular reference solving')

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        sys.exit(0)

    commands = {
        'calc': cmd_calc,
        'build': cmd_build,
        'convert': cmd_convert,
        'parse': cmd_parse,
        'info': cmd_info,
        'sheets': cmd_sheets,
        'serve': cmd_serve,
    }

    commands[args.command](args)


if __name__ == '__main__':
    main()
